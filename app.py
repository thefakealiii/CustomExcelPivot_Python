from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter
from openpyxl import Workbook
import pandas as pd
import os

#Constant Variables
monthSetting = "2021-11-"

#Array to handle other value
sizesChangeArr = ['35', '36', '37', '38', '39', '40', '41', '42', '43', '44', '45', '46', '47', '48', '49', '50']
#Array to run the piovt	
sizesArr = ['35', '36', '37', '38', '39', '40', '41', '42', '43', '44', '45', '46', '47', '48', '49', '50', 'Others','Aggregated']

rowww = 0

#Parent loop to traverse through all the dates
for i in range(1,31):
	try:
		#Open the existing workbook
		wb = load_workbook(monthSetting+str(i)+'.xlsx')
		ws = wb['Sheet1']
		#Data Initialization
		testCount = 0
		rowSize = ws.max_row
		colSize = ws.max_column
		colSize = colSize+1
		#Dictionary Array to save all the values
		dictArray = [{'Date':monthSetting+str(i), 'Size': sizesArr[k],'Qty':0,'Total':0} for k in range(len(sizesArr))]
		#Loop to give 'Other' value to cells where required
		for ii in range(rowSize-1):
			check = 0
			for jj in range(len(sizesChangeArr)):
				if ws['H'+str(ii+2)].value == None:
					check += 1
				elif ws['H'+str(ii+2)].value.__contains__(sizesChangeArr[jj]) == False:
					check += 1
			if check == len(sizesChangeArr):
				ws['H'+str(ii+2)] = 'Others'
		#Save the updated file of Others value in a temporary file
		wb.save(monthSetting+str(i)+'temp.xlsx')
		#Load the temporary file to perform operations
		wb1 = load_workbook(monthSetting+str(i)+'temp.xlsx')
		ws1 = wb1['Sheet1']
		totalPrice = 0
		#Perform the required operations of the algorithm
		for ia in range(rowSize-1):
			for ja in range(len(sizesArr)-1):
				if ws['H'+str(ia+2)].value.__contains__(sizesArr[ja]):
					testCount += 1
					dictArray[ja]["Qty"] += 1
					dictArray[ja]["Total"] += float(ws['F'+str(ia+2)].value)	

			totalPrice = float(ws['F'+str(ia+2)].value) + totalPrice
		indx = len(sizesArr) - 1
		dictArray[indx]["Qty"] = int(testCount)
		dictArray[indx]["Total"] = float(totalPrice)

		#Read the output file and save the data there
	
		writer = pd.ExcelWriter("result.xlsx", mode="a", if_sheet_exists='overlay', engine = 'openpyxl')
		df=pd.DataFrame(dictArray)
		df = df.transpose()

		df.to_excel(writer, startrow = rowww,index = False)
		rowww = rowww + 5

		writer.close()

		#Delete the temporary file to save disk space
		os.remove(monthSetting+str(i)+'temp.xlsx')

	#If a file is not readble rather than crashing the application show an error on the console
	except IOError: 
		print(monthSetting+str(i)+'.xlsx does not exist.')

